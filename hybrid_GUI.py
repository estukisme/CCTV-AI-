import cv2
# Force FFMPEG use HEVC (H.265)
import os
cv2.setNumThreads(1)  # stabilizer multi-thread decoder
os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = (
    "video_codec;h264;"
    "rtsp_transport;tcp;"
    "allowed_media_types;video;"
    "analyzeduration;0;"
    "probesize;32768;"
    "fflags;discardcorrupt;"
)
import numpy as np
import tensorflow as tf
import torch
from ultralytics import YOLO
import tkinter as tk
from tkinter import Label, ttk
from tkinter.scrolledtext import ScrolledText
from PIL import Image, ImageTk
from datetime import datetime
import threading
import winsound
import time
def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import openpyxl
import random
import tensorflow as tf
print(tf.config.list_physical_devices('GPU'))

# =========================================================
# GPU SETUP
# =========================================================
torch.backends.cudnn.benchmark = True
DEVICE = "cuda"

print("Torch:", torch.__version__)
print("CUDA:", torch.cuda.is_available())
print("GPU:", torch.cuda.get_device_name(0))

gpus = tf.config.list_physical_devices('GPU')
if gpus:
    for gpu in gpus:
        tf.config.experimental.set_memory_growth(gpu, True)

# =========================================================
# GLOBAL
# =========================================================
running = False
processing = False
alarm_active = False
buzzer_running = False
current_frame = None
last_time = time.time()
fps = 0
export_enabled = True
current_shift_name = None
excel_lock = threading.Lock()
ROI_X1 = 500
ROI_Y1 = 200
ROI_X2 = 1400
ROI_Y2 = 850
current_shift_alarm_file = None
current_shift_monitor_file = None
current_shift_bongkahan_file = None
last_monitor_minute = None
last_alarm_time = 0
frame_valid = True
rtsp_url = (
    "rtsp://admin:Phonska1bisa@192.168.110.3:554/"
    "Streaming/Channels/102"
    "?rtsp_transport=tcp"
    "&enableHevc=0"
    "&fflags=discardcorrupt"
    "&flags=low_delay"
    "&probe_size=128000"
    "&analyzeduration=0"
    "&reorder_queue_size=5"
)
display_mode = "heatmap"   # heatmap | texture | off
# Auto reconnect timer
AUTO_RESTART_BASE = 6 * 3600             # 6 jam
AUTO_RESTART_OFFSET = random.randint(180, 360)  # 3–6 menit acak
last_global_restart = time.time()
rtsp_lost_counter = 0
last_status = "-"
status_hold_frames = 0
STATUS_HOLD_LIMIT = 8
last_processed_frame = None


# =========================================================
# LOAD MODELS
# =========================================================
yolo = YOLO("yolov8n.pt")
yolo.to(DEVICE)
yolo.fuse()

# =========================================================
# TENSORFLOW GPU FORCE ENABLE (WAJIB)
# =========================================================
gpus = tf.config.list_physical_devices('GPU')
if gpus:
    try:
        tf.config.set_visible_devices(gpus[0], 'GPU')
        tf.config.experimental.set_memory_growth(gpus[0], True)
        print("TensorFlow menggunakan GPU:", gpus[0])
    except Exception as e:
        print("TF GPU ERROR:", e)
else:
    print("Tidak ada GPU untuk TensorFlow!")

with tf.device('/GPU:0'):
    cnn = tf.keras.models.load_model(
    "cnn/model_cnn_v2.h5",
    compile=False
)

cnn_labels = ["merah", "tidak_merah"]
torch.set_float32_matmul_precision('high')

# =========================================================
# SHIFT
# =========================================================
def get_shift(now):
    hour = now.hour
    if 7 <= hour < 15:
        return "Shift 1 (07:00-15:00)"
    elif 15 <= hour < 23:
        return "Shift 2 (15:00-23:00)"
    else:
        return "Shift 3 (23:00-07:00)"
    
def get_alarm_filename(now):
    shift = get_shift(now)
    date_str = now.strftime("%Y-%m-%d")
    return f"{date_str}_{shift.replace(' ','_').replace(':','')}_ALARM.xlsx"

def get_monitor_filename(now):
    shift = get_shift(now)
    date_str = now.strftime("%Y-%m-%d")
    return f"{date_str}_{shift.replace(' ','_').replace(':','')}_MONITORING.xlsx"
def get_bongkahan_filename(now):
    shift = get_shift(now)
    date_str = now.strftime("%Y-%m-%d")
    return f"{date_str}_{shift.replace(' ','_').replace(':','')}_BONGKAHAN.xlsx"
import requests

TOKEN = "8594979826:AAGo7-o9ELAnmg5qR-pwM5HYObKwvFmG168"
CHAT_IDS = ["8141547822", "8072941565"]

def kirim_tele_gambar(path, caption=""):
    url = f"https://api.telegram.org/bot{TOKEN}/sendPhoto"

    if not os.path.exists(path):
        log(f"TELEGRAM ❌ File tidak ditemukan: {path}")
        return

    if os.path.getsize(path) < 2000:
        log(f"TELEGRAM ❌ File rusak/terlalu kecil: {path}")
        return

    for chat_id in CHAT_IDS:
        try:
            with open(path, "rb") as f:
                r = requests.post(url, files={"photo": f}, data={
                    "chat_id": chat_id,
                    "caption": caption
                }, timeout=12)

            if '"ok":true' in r.text:
                log(f"TELEGRAM ✔ Alarm terkirim → {chat_id}")
            else:
                log(f"TELEGRAM ❌ Gagal kirim → {chat_id}")

        except Exception as e:
            log(f"TELEGRAM ⚠ ERROR → {chat_id}: {e}")   
    
# =========================================================
# EXPORT EXCEL
# =========================================================
def append_to_excel(data,filename):


    if filename is None:
        return

    with excel_lock:

        os.makedirs("export", exist_ok=True)
        filepath = os.path.join("export", filename)

        if os.path.exists(filepath):
            try:
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
            except:
                os.remove(filepath)
                wb = Workbook()
                ws = wb.active
                ws.title = "Log Produk"
                ws.append(["Foto Produk", "Tanggal", "Jam", "Shift", "Warna Produk"])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log Produk"
            ws.append(["Foto Produk", "Tanggal", "Jam", "Shift", "Warna Produk"])
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 18

        row = ws.max_row + 1

        ws.cell(row=row, column=2, value=data["date"])
        ws.cell(row=row, column=3, value=data["time"])
        ws.cell(row=row, column=4, value=data["shift"])
        ws.cell(row=row, column=5, value=data["status"])

        if os.path.exists(data["image_path"]):
            img = XLImage(data["image_path"])
            img.width = 120
            img.height = 120
            ws.add_image(img, f"A{row}")
            ws.row_dimensions[row].height = 95

        temp_path = filepath + ".tmp"
        wb.save(temp_path)
        wb.close()
        # Replace file lama secara aman
        os.replace(temp_path, filepath)
    
# =========================================================
# BUZZER
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOUND_FILE = os.path.join(BASE_DIR, "alarm.wav")
def buzzer_loop():
    global buzzer_running

    while buzzer_running:

        # Play sampai selesai (blocking, bukan async)
        winsound.PlaySound(SOUND_FILE, winsound.SND_FILENAME | winsound.SND_ASYNC)

        # Setelah suara selesai, jeda 3 detik sebelum mengulang
        time.sleep(3)   

def start_buzzer():
    global buzzer_running
    if not buzzer_running:
        buzzer_running = True
        threading.Thread(target=buzzer_loop, daemon=True).start()

def stop_buzzer():
    global buzzer_running
    buzzer_running = False
    winsound.PlaySound(None, winsound.SND_PURGE)

def resize_with_aspect_ratio(image, width=None, height=None):
    h, w = image.shape[:2]

    if width is None and height is None:
        return image

    if width is not None:
        ratio = width / w
        new_dim = (width, int(h * ratio))
    else:
        ratio = height / h
        new_dim = (int(w * ratio), height)

    return cv2.resize(image, new_dim)

def toggle_texture():
    global texture_enabled
    texture_enabled = not texture_enabled
    btn_texture.config(text=f"Texture: {'ON' if texture_enabled else 'OFF'}")
   
# =========================================================
# CNN PREPROCESS
# =========================================================
def preprocess_cnn(image_bgr):
    rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    resized = cv2.resize(rgb, (64, 64))
    return resized.astype("float32") / 255.0

# =========================================================
# DETECTION
# =========================================================
def detect_objects(frame):
    
    global alarm_active, processing, current_frame
    global current_shift_name
    global ROI_X1, ROI_Y1, ROI_X2, ROI_Y2
    global current_shift_alarm_file, current_shift_monitor_file
    global last_monitor_minute, export_enabled,last_alarm_time
    global frame_valid,current_shift_bongkahan_file
    global last_status, status_hold_frames
    global last_processed_frame
    #cv2.rectangle(frame,(ROI_X1, ROI_Y1),(ROI_X2, ROI_Y2),(200,200,200),2)
    frame_valid = True
    detected_color = "-"
    status = "-"
    now_t = time.time()
    
    # ========== BLOCK BAD FRAMES (GLOBAL FRAME) ==========         
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    brightness = gray.mean()
    variance = gray.var()

    bad_frame = False

    # 1) Terlalu gelap / noisy
    if brightness < 35:
        bad_frame = True

    if variance < 40:
        bad_frame = True

    # 2) Frame korup (freeze H.265)
    if np.max(frame) < 50:
        bad_frame = True
    if np.mean(frame) < 15:
        bad_frame = True

    # 3) ROI invalid
    roi = frame[ROI_Y1:ROI_Y2, ROI_X1:ROI_X2]
    if roi.size == 0:
        bad_frame = True
    if roi.shape[0] < 50 or roi.shape[1] < 50:
        bad_frame = True

    # === STOP DI SINI kalau frame buruk ===
    if bad_frame:
        frame_valid = False
        detected_color = "-"
        processing = False
        current_frame = frame
        return

    # ======================================================
    # =======================================
    detected_color = "-"
    roi = frame[ROI_Y1:ROI_Y2, ROI_X1:ROI_X2]

    # ======================================================
    #       HARD BLOCK BAD FRAME SUPAYA TIDAK MASUK CNN
    # ======================================================
    # 1. ROI kosong / kecil
    if roi is None or roi.size == 0 or roi.shape[0] < 60 or roi.shape[1] < 60:
        frame_valid = False
        detected_color = "-"
        processing = False
        current_frame = frame
        return

    # 2. Gelap / overexposure / ghost green / pink
    roi_mean = roi.mean()
    roi_max  = roi.max()
    roi_var  = roi.var()

    if roi_mean < 40 or roi_max < 70 or roi_var < 50:
        frame_valid = False
        detected_color = "-"
        processing = False
        current_frame = frame
        return

    # 3. Deteksi ghost frame HSV (paling penting!)
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    h = hsv[:,:,0].mean()
    s = hsv[:,:,1].mean()

    # ghost = saturasi tinggi + hue abnormal (pink/hijau)
    if (h > 140 and s > 60) or (h < 10 and s > 70):
        frame_valid = False
        detected_color = "-"
        processing = False
        current_frame = frame
        return
    
    # ========== FILTER ANTI IR / ANTI RTSP LOST ==========
    roi_hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    h_mean = roi_hsv[:, :, 0].mean()
    s_mean = roi_hsv[:, :, 1].mean()
    v_mean = roi_hsv[:, :, 2].mean()
    IR_MODE = False
    # frame IR / ghost grey / freeze
    if (140 <= h_mean <= 185) and (s_mean >= 40):
        IR_MODE = True
# =======================================================
    if roi.size != 0:

        # === ROI Center Small (SUPER STABIL) ===

        roi = frame[ROI_Y1:ROI_Y2, ROI_X1:ROI_X2]

        h, w = roi.shape[:2]

        # hanya ambil 25% area tengah
        cw = int(w * 0.85)
        ch = int(h * 0.85)

        x1 = (w - cw) // 2
        y1 = (h - ch) // 2
        x2 = x1 + cw
        y2 = y1 + ch

        roi_center = roi[y1:y2, x1:x2]
        
        lab = cv2.cvtColor(roi_center, cv2.COLOR_BGR2LAB)
        L, A, B = cv2.split(lab)

        # Persempit range L agar warna tidak berubah karena exposure
        clahe = cv2.createCLAHE(clipLimit=1.2, tileGridSize=(8,8))
        L = clahe.apply(L)
        
        # ===== IR COLOR FIX =====
        if IR_MODE:
            # IR menghasilkan pink → netralisasi ke LAB abu-abu
            A[:] = 128
            B[:] = 128
            lab_ir = cv2.merge((L, A, B))
            roi_fixed = cv2.cvtColor(lab_ir, cv2.COLOR_LAB2BGR)
        else:
            lab_eq = cv2.merge((L, A, B))
            roi_fixed = cv2.cvtColor(lab_eq, cv2.COLOR_LAB2BGR)
        # ===============================
        # DETEKSI CONVEYOR KOSONG
        # ===============================

        gray_roi = cv2.cvtColor(roi_center, cv2.COLOR_BGR2GRAY)
        gray_roi = cv2.medianBlur(gray_roi,5)
        texture = gray_roi.var()

        # jika tekstur terlalu kecil → tidak ada produk
        edges = cv2.Canny(gray_roi,80,160)
        edge_count = np.count_nonzero(edges)
        # ===============================
        # DETEKSI BUTIRAN PUPUK
        # ===============================
        _, grain_mask = cv2.threshold(gray_roi, 70, 255, cv2.THRESH_BINARY)

        
        # ===============================
        # CEK ADA PRODUK ATAU TIDAK
        # ===============================
        grain_pixels = np.count_nonzero(grain_mask)
        grain_ratio = grain_pixels / gray_roi.size

        # produk terlalu sedikit → NO PRODUCT
        if grain_ratio < 0.12:
            status = last_status
            detected_color = last_status
            text = "NO PRODUCT"
            font = cv2.FONT_HERSHEY_SIMPLEX
            scale = 2.2
            thickness = 5
            (text_w, text_h), _ = cv2.getTextSize(text, font, scale, thickness)
            frame_h, frame_w = frame.shape[:2]
            x = (frame_w - text_w) // 2
            y = int(frame_h * 0.15)
            cv2.putText(frame,text,(x,y),
                font,scale,(0,0,0),thickness+6,cv2.LINE_AA)
            cv2.putText(frame,text,(x,y),
                font,scale,(255,255,255),thickness,cv2.LINE_AA)
            processing = False
            current_frame = frame
            return
        bright_pixels = np.sum(gray_roi > 150)
        density = bright_pixels / gray_roi.size
        # ===============================
        # DETEKSI BELT KOSONG (STABIL)
        # ===============================
        if density < 0.02 and (texture < 20 and edge_count < 20):
            status = "-"
            detected_color = "-"

            text = "NO PRODUCT"

            font = cv2.FONT_HERSHEY_SIMPLEX
            scale = 2.2
            thickness = 5

            (text_w, text_h), _ = cv2.getTextSize(text, font, scale, thickness)

            frame_h, frame_w = frame.shape[:2]

            x = (frame_w - text_w) // 2
            y = int(frame_h * 0.15)

            cv2.putText(frame,text,(x,y),
                font,scale,(0,0,0),thickness+6,cv2.LINE_AA)

            cv2.putText(frame,text,(x,y),
                font,scale,(255,255,255),thickness,cv2.LINE_AA)

            processing = False
            current_frame = frame
            return
        belt_hsv = cv2.cvtColor(roi_center, cv2.COLOR_BGR2HSV)

        belt_mask = cv2.inRange(
            belt_hsv,
            (10, 20, 20),   # lower
            (18, 120, 120) # upper
        )

        belt_ratio = np.count_nonzero(belt_mask) / belt_mask.size

        if belt_ratio > 0.90:
            status = "-"
            detected_color = "-"
            processing = False
            current_frame = frame
            return
    # ===============================
    # PRODUK TERLALU SEDIKIT → NO PRODUCT
    # ===============================
        skip_cnn = False
        if grain_ratio < 0.35:
            skip_cnn = True
        if not skip_cnn:
            # kirim ke CNN
            inp = preprocess_cnn(roi_fixed).reshape(1, 64, 64, 3)
            inp = inp.astype(np.float32)
            inp = np.ascontiguousarray(inp)
            pred = cnn(inp, training=False).numpy()[0]

            confidence_merah = pred[0]
            confidence_tidak = pred[1]

            if confidence_merah > confidence_tidak + 0.10:
                detected_color = "merah"
            else:
                detected_color = "tidak_merah"
            status = detected_color
        else:
            # produk terlalu sedikit → pertahankan status lama
            status = last_status
        # inisialisasi pertama
        if last_status == "-"  and status != "-":
            last_status = status
        # ===============================
        # STATUS STABILIZER (ANTI FLICKER)
        # ===============================

        if status == "-" or status is None:
            status = last_status
            status_hold_frames = 0

        elif status == last_status:
            status_hold_frames = 0
        else:
            status_hold_frames += 1

            if status_hold_frames >= STATUS_HOLD_LIMIT:
                last_status = status
                status_hold_frames = 0
            else:
                status = last_status
        if status == "merah":
            color_box = (0,0,255)

        elif status == "tidak_merah":
            color_box = (0,255,255)

        elif status == "bongkahan":
            color_box = (0,165,255)

        cv2.rectangle(frame,(ROI_X1, ROI_Y1),(ROI_X2, ROI_Y2),color_box,3)
        
        cv2.putText(frame,status.upper(),
                    (ROI_X1, ROI_Y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,color_box,3)
       # Proteksi terakhir — tidak boleh export bad frame
        if  detected_color == "-" or roi_center is None or roi_center.size == 0:
            frame_valid = False
            processing = False
            current_frame = frame
            return
            
        if export_enabled:

            now = datetime.now()
            shift_name = get_shift(now)

            # ===== SHIFT CHANGE =====
            if current_shift_name != shift_name:
                current_shift_name = shift_name
                current_shift_alarm_file = get_alarm_filename(now)
                current_shift_monitor_file = get_monitor_filename(now)
                current_shift_bongkahan_file = get_bongkahan_filename(now)
                last_monitor_minute = None

            # ===== ALARM FILE =====
            if (
                frame_valid
                and status == "tidak_merah"
                and roi_center is not None
                and roi_center.size > 0
                and roi_center.mean() > 55
                and roi_center.var() > 60
                and (now_t - last_alarm_time > 30)
            ):
                last_alarm_time = now_t  # update cooldown
                
                os.makedirs("snapshot_alarm", exist_ok=True)
                img_name = now.strftime("%Y%m%d_%H%M%S") + "_ALARM.jpg"
                img_path = os.path.join("snapshot_alarm", img_name)
                cv2.imwrite(img_path, roi_center)
                

                # SETELAH gambar siap → kirim ke Telegram
                threading.Thread(
                target=kirim_tele_gambar,
                args=(
                    img_path,
                    f"⚠️ ALARM CCTV!\nStatus: {status.upper()}\nWaktu: {now.strftime('%H:%M:%S')}"
                ),
                daemon=True
                ).start()

                data = {
                    "date": now.strftime("%Y-%m-%d"),
                    "time": now.strftime("%H:%M:%S"),
                    "shift": shift_name,
                    "status": status,
                    "image_path": img_path
                }

                append_to_excel(data, current_shift_alarm_file)
            # ===== BONGKAHAN FILE =====
            if (
                frame_valid
                and status == "bongkahan"
                and roi_center is not None
                and roi_center.size > 0
                and roi_center.mean() > 55
                and roi_center.var() > 60
                and (now_t - last_alarm_time > 30)
            ):

                last_alarm_time = now_t

                os.makedirs("snapshot_bongkahan", exist_ok=True)

                img_name = now.strftime("%Y%m%d_%H%M%S") + "_BONGKAHAN.jpg"
                img_path = os.path.join("snapshot_bongkahan", img_name)

                cv2.imwrite(img_path, roi_center)

                kirim_tele_gambar(
                    img_path,
                    f"⚠️ ALARM BONGKAHAN!\nStatus: {status.upper()}\nWaktu: {now.strftime('%H:%M:%S')}"
                )

                data = {
                    "date": now.strftime("%Y-%m-%d"),
                    "time": now.strftime("%H:%M:%S"),
                    "shift": shift_name,
                    "status": status,
                    "image_path": img_path
                }
                append_to_excel(data, current_shift_bongkahan_file)

            # ===== MONITORING FILE (00 & 30) =====
            if now.minute in [0, 30] and now.second <= 5:

                if last_monitor_minute != now.minute:

                    last_monitor_minute = now.minute

                    os.makedirs("snapshot_monitoring", exist_ok=True)
                    img_name = now.strftime("%Y%m%d_%H%M%S") + ".jpg"
                    img_path = os.path.join("snapshot_monitoring", img_name)
                    cv2.imwrite(img_path, roi_center)

                    data = {
                        "date": now.strftime("%Y-%m-%d"),
                        "time": now.strftime("%H:%M:%S"),
                        "shift": shift_name,
                        "status": status,
                        "image_path": img_path
                    }

                    append_to_excel(data, current_shift_monitor_file)
                    

    # ===== BUZZER CONTROL =====
    if frame_valid and status in ["tidak_merah","bongkahan"]:
        if not alarm_active:
            alarm_active = True
            start_buzzer()
    else:
        if alarm_active:
            alarm_active = False
            stop_buzzer()

    current_frame = frame.copy()
    last_processed_frame = current_frame
    processing = False


# =========================================================
# HEATMAP
# =========================================================

def create_heatmap(frame):
    # Ambil channel merah
    red_channel = frame[:, :, 2]

    # Normalize
    heat = cv2.normalize(red_channel, None, 0, 255, cv2.NORM_MINMAX)

    # Convert jadi colormap
    heatmap = cv2.applyColorMap(heat, cv2.COLORMAP_JET)

    # Perkecil heatmap
    heatmap_small = cv2.resize(heatmap, (220, 150))

    return heatmap_small

def create_texture_map(frame):
    # Grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # =====================================================
    # 1. Super High-Pass (angkat detail ekstrem)
    # =====================================================
    blur = cv2.GaussianBlur(gray, (31, 31), 0)
    high_pass = cv2.addWeighted(gray, 2.5, blur, -1.5, 0)

    # =====================================================
    # 2. Multi-directional Emboss (8 arah)
    # =====================================================
    kernels = [
        np.array([[-1, -1,  0],
                  [-1,  0,  1],
                  [ 0,  1,  1]], dtype=np.float32),

        np.array([[ 0, -1, -1],
                  [ 1,  0, -1],
                  [ 1,  1,  0]], dtype=np.float32),

        np.array([[-1,  0,  1],
                  [-1,  0,  1],
                  [-1,  0,  1]], dtype=np.float32),

        np.array([[ 1,  1,  1],
                  [ 0,  0,  0],
                  [-1,-1,-1]], dtype=np.float32)
    ]

    emboss_sum = np.zeros_like(high_pass, dtype=np.float32)

    for k in kernels:
        emboss_sum += cv2.filter2D(high_pass, -1, k)

    emboss_sum = cv2.normalize(emboss_sum, None, 0, 255, cv2.NORM_MINMAX).astype("uint8")

    # =====================================================
    # 3. CLAHE (perjelas kedalaman permukaan)
    # =====================================================
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(emboss_sum)

    # =====================================================
    # 4. Sharpen + Depth Boost
    # =====================================================
    sharpen_kernel = np.array([
        [ 0, -1,  0],
        [-1,  6, -1],
        [ 0, -1,  0]
    ], dtype=np.float32)

    deep = cv2.filter2D(enhanced, -1, sharpen_kernel)

    # =====================================================
    # 5. Final depth compression (biar halus tapi tetap timbul)
    # =====================================================
    final = cv2.normalize(deep, None, 0, 255, cv2.NORM_MINMAX)
    final = final.astype("uint8")

    # Convert ke BGR dan kecilkan
    final_bgr = cv2.cvtColor(final, cv2.COLOR_GRAY2BGR)
    final_small = cv2.resize(final_bgr, (220, 150))

    return final_small

def toggle_mode():
    global display_mode

    if display_mode == "heatmap":
        display_mode = "texture"
        btn_mode.config(text="Mode: TEXTURE")
    elif display_mode == "texture":
        display_mode = "off"
        btn_mode.config(text="Mode: OFF")
    else:
        display_mode = "heatmap"
        btn_mode.config(text="Mode: HEATMAP")



def force_reconnect():
    global cap,processing
    
    log("FORCE RECONNECTING RTSP ...")
    processing = False  # stop semua deteksi
    
    try:
        if cap is not None:
            cap.release()
    except:
        pass
    time.sleep(0.5)

     # Jangan pakai D3D11 saat reconnect
    os.environ["OPENCV_FFMPEG_HW_ACCELERATION"] = "d3d11va"

    cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
    
    
     # -- SAFE CHECK --
    if not cap.isOpened():
        log("RTSP NOT OPENED!")
        return False

    # Safe backend check
    try:
        log(f"Backend: {cap.getBackendName()}")
    except:
        log("Backend: UNKNOWN")

    cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 2000)
    cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 2000)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
    cap.set(cv2.CAP_PROP_FPS, 30)

    return True

# =========================================================
# MAIN LOOP (STABIL)
# =========================================================
def main_loop():
    global running, processing, current_frame
    global last_time, fps
    global cap,rtsp_lost_counter
    global ROI_X1, ROI_Y1, ROI_X2, ROI_Y2
    global last_processed_frame
    
    if not running:
        return
    # =========================================
    # AUTO RECONNECT setiap 6 jam + offset 3–6 menit
    # =========================================
    global last_global_restart, AUTO_RESTART_OFFSET

    if time.time() - last_global_restart > (AUTO_RESTART_BASE + AUTO_RESTART_OFFSET):
        log(f"[AUTO] Restart RTSP client after 6h+{AUTO_RESTART_OFFSET}s")
        rtsp_status_label.config(text="AUTO RESTARTING...", fg="yellow")

        force_reconnect()

        # Reset timer + buat offset acak baru
        last_global_restart = time.time()
        AUTO_RESTART_OFFSET = random.randint(180, 360)
        
        # Penting: hentikan loop frame ini
        root.after(50, main_loop)
        return
    ret, frame = cap.read()
    if frame is not None:
        if frame.std() < 1:
            print("[DROP] Possible frozen frame")
    
    if not ret or frame is None:
        rtsp_lost_counter += 1

        if rtsp_lost_counter > 10:
            rtsp_status_label.config(text="RTSP: LOST", fg="red")
        log("RTSP frame lost")
        root.after(20, main_loop)
        return
    else:
        rtsp_lost_counter = 0
        rtsp_status_label.config(text="RTSP: CONNECTED", fg="lime")
    # === ROI FIX FINAL UNTUK SUBSTREAM 1280×720 ===
    FRAME_W = frame.shape[1]
    FRAME_H = frame.shape[0]

    # === ROI FIX FINAL UNTUK SUBSTREAM 1280×720 ===
    ROI_X1 = 300
    ROI_Y1 = 80
    ROI_X2 = 985
    ROI_Y2 = 715
    
    # === RTSP STATUS UPDATE ===
    if not ret or frame is None:
        # LOST sebelum proses deeper
        rtsp_status_label.config(text="RTSP: LOST — Reconnecting...", fg="yellow")
    else:
        rtsp_status_label.config(text="RTSP: CONNECTED", fg="lime")

    # FIX H.265 blurring / corrupt partial frames
    if ret and (frame is None or frame.size < 120000):
        print("[DROP] H.265 bad frame (POC missing), skipping...")
        rtsp_status_label.config(text="RTSP: BAD FRAME (H265)", fg="red")

        ok = force_reconnect()
        if not ok:
            print("Reconnect failed, retrying...")
            rtsp_status_label.config(text="RTSP: FAILED", fg="red")
            root.after(1500, main_loop)  # tidak spam reconnect, CPU aman
            return

        rtsp_status_label.config(text="RTSP: CONNECTED", fg="lime")
        root.after(50, main_loop)
        return

    # === GLOBAL BAD FRAME FILTER (LEVEL 1) ===
    if frame.size < 120000 or np.max(frame) < 30:
        rtsp_status_label.config(text="RTSP: BAD FRAME SKIP", fg="red")
        root.after(1, main_loop)
        return

    # Jalankan deteksi async
    if not processing:
        processing = True
    threading.Thread(
        target=detect_objects,
        args=(frame.copy(),),
        daemon=True
    ).start()
    
    # Frame display
    if last_processed_frame is not None:
        display = last_processed_frame.copy()
    else:
        display = frame.copy()
    #cv2.rectangle(display,(ROI_X1, ROI_Y1),(ROI_X2, ROI_Y2),(180,180,180),2)

    root.update_idletasks()
    container_w = video_container.winfo_width()
    container_h = video_container.winfo_height()

    if container_w > 10 and container_h > 10:
        h, w = display.shape[:2]
        scale = min(container_w / w, container_h / h)
        new_w = int(w * scale)
        new_h = int(h * scale)
        resized = cv2.resize(display, (new_w, new_h))
        # Buat canvas hitam sebesar container
        canvas = np.zeros((container_h, container_w, 3), dtype=np.uint8)

        # Hitung offset tengah
        x_offset = (container_w - new_w) // 2
        y_offset = 0

        # Tempel video di tengah
        canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized

        display = canvas

    # ==========================
    # FPS CALCULATION
    # ==========================
    now = time.time()
    fps = 1 / (now - last_time)
    last_time = now

    fps_int = int(fps)
    color = "lime" if fps_int >= 8 else ("yellow" if fps_int >= 3 else "red")
    fps_overlay.config(text=f"FPS: {fps_int}", fg=color)

    # ==========================
    # DISPLAY MODE (1 PILIHAN)
    # ==========================
    if display_mode == "heatmap":
        heatmap = create_heatmap(display)
        hm_rgb = cv2.cvtColor(heatmap, cv2.COLOR_BGR2RGB)
        hm_img = ImageTk.PhotoImage(Image.fromarray(hm_rgb))
        heatmap_label.imgtk = hm_img
        heatmap_label.configure(image=hm_img)

    elif display_mode == "texture":
        texture = create_texture_map(display)
        tex_rgb = cv2.cvtColor(texture, cv2.COLOR_BGR2RGB)
        tex_img = ImageTk.PhotoImage(Image.fromarray(tex_rgb))
        heatmap_label.imgtk = tex_img
        heatmap_label.configure(image=tex_img)

    else:
        blank = np.zeros((150, 220, 3), dtype=np.uint8)
        blank_rgb = cv2.cvtColor(blank, cv2.COLOR_BGR2RGB)
        blank_img = ImageTk.PhotoImage(Image.fromarray(blank_rgb))
        heatmap_label.imgtk = blank_img
        heatmap_label.configure(image=blank_img)
        
    # ==========================
    # VIDEO DISPLAY
    # ==========================
    rgb = cv2.cvtColor(display, cv2.COLOR_BGR2RGB)
    img = ImageTk.PhotoImage(Image.fromarray(rgb))

    video_label.imgtk = img
    video_label.configure(image=img)
    
    # FPS stabilizer ideal around 25 FPS
    TARGET_FPS = 35
    FRAME_DELAY = int(1000 / TARGET_FPS)
    root.after(FRAME_DELAY, main_loop)


# =========================================================
# SNAPSHOT
# =========================================================
def snapshot():
    if current_frame is None:
        return
    fname = datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"
    cv2.imwrite(os.path.join("snapshot", fname), current_frame)

def toggle_export():
    global export_enabled
    export_enabled = not export_enabled
    btn_export.config(text=f"Export: {'ON' if export_enabled else 'OFF'}")
# =========================================================
# CONTROL
# =========================================================
def start_camera():
    global running
    running = True
    dashboard_frame.pack_forget()
    camera_frame.pack(fill="both", expand=True)
    main_loop()

def stop_camera():
    global running
    running = False
    stop_buzzer()

def back_to_dashboard():
    global running
    running = False
    stop_buzzer()
    camera_frame.pack_forget()
    dashboard_frame.pack(fill="both", expand=True)

# =========================================================
# GUI
# =========================================================
root = tk.Tk()
root.title("Hybrid YOLO + CNN RTX 4070 Ti STABLE")
root.state("zoomed")
root.configure(bg="#1e1e1e")

style = ttk.Style()
style.theme_use("default")

# ==========================
# DASHBOARD
# ==========================
dashboard_frame = tk.Frame(root, bg="#1e1e1e")

tk.Label(dashboard_frame, text="Dashboard",
         fg="white", bg="#1e1e1e",
         font=("Arial", 26, "bold")).pack(pady=50)

ttk.Button(dashboard_frame, text="Mulai Kamera",
           command=start_camera).pack(pady=20)

ttk.Button(dashboard_frame, text="Keluar",
           command=root.quit).pack(pady=20)

dashboard_frame.pack(fill="both", expand=True)

# ==========================
# CAMERA FRAME
# ==========================

camera_frame = tk.Frame(root, bg="#1e1e1e")

# ==========================
# TOP BAR MONITORING (ANTI GESER & ANTI NUMPUK)
# ==========================
top_bar = tk.Frame(camera_frame, bg="#0f0f0f", height=60)
top_bar.pack(fill="x")

# 3 kolom stabil
top_bar.grid_columnconfigure(0, weight=1)
top_bar.grid_columnconfigure(1, weight=1)
top_bar.grid_columnconfigure(2, weight=1)

# === LEFT LABEL ===
title_label = tk.Label(
    top_bar,
    text="CCTV AI PRODUK M-118",
    font=("Arial", 20, "bold"),
    bg="#0f0f0f",
    fg="#00A8FF"
)
title_label.grid(row=0, column=0, sticky="w", padx=30)


# === RIGHT CONTAINER (LOGO + FPS) ===
right_container = tk.Frame(top_bar, bg="#0f0f0f")
right_container.grid(row=0, column=2, sticky="e", padx=15)

# Logo loader
def load_logo(path):
    img = Image.open(path)
    base_height = 55
    ratio = base_height / img.size[1]
    img = img.resize((int(img.size[0]*ratio), base_height), Image.LANCZOS)
    return ImageTk.PhotoImage(img)

# Logo paths
logo_paths = [
    "logo.png",
    "LOGO 2A.png",
    "logo KIK new 2.png",
    "POG.jpg",
    "Logo Danantara.png",
    "Logo Pupuk Indonesia.png"
]

for p in logo_paths:
    try:
        tk_img = load_logo(p)
        lbl = tk.Label(right_container, image=tk_img, bg="#ffffff")
        lbl.image = tk_img
        lbl.pack(side="left", padx=10)
    except:
        print("Gagal load logo:", p)
# ==========================
# CONTENT FRAME (CENTER MODE)
# ==========================
content_frame = tk.Frame(camera_frame, bg="#1e1e1e")
content_frame.pack(fill="both", expand=True)


# 1 kolom saja
content_frame.columnconfigure(0, weight=0)
content_frame.columnconfigure(1, weight=1)  # video fleksibel

# 3 row
content_frame.rowconfigure(0, weight=1)  # JUDUL kecil

# ==========================
# PANEL KIRI (BUTTON)
# ==========================
btn_panel = tk.Frame(content_frame, bg="#1e1e1e")
btn_panel.grid(row=0, column=0, sticky="ns", padx=20, pady=20)

ttk.Button(btn_panel, text="Stop",
           command=stop_camera).pack(fill="x", pady=(0, 15))

ttk.Button(btn_panel, text="Snapshot",
           command=snapshot).pack(fill="x", pady=(0, 15))

btn_mode = ttk.Button(btn_panel,
                      text="Mode: HEATMAP",
                      command=toggle_mode)
btn_mode.pack(fill="x", pady=(0, 15))

btn_export = ttk.Button(btn_panel,
                        text="Export: ON",
                        command=toggle_export)
btn_export.pack(fill="x", pady=(0, 15))


ttk.Button(btn_panel, text="Kembali",
           command=back_to_dashboard).pack(fill="x")

heatmap_label = tk.Label(btn_panel, bg="black")
heatmap_label.pack(pady=(30, 0))

# === LEFT TITLE (2 WARNA) ===
title_frame = tk.Frame(btn_panel, bg="#0f0f0f")
title_frame.pack(pady=(30,0))

title1 = tk.Label(
    title_frame,
    text="DEP. OPERASI PABRIK IIA",
    font=("Segoe UI", 18, "bold"),
    bg="#0f0f0f",
    fg="#00ae4d"   
)
title1.pack()

line = tk.Label(
    title_frame,
    text="=====================",
    font=("Consolas", 12, "bold"),
    bg="#1e1e1e",
    fg="white"
)
line.pack()

title2 = tk.Label(
    title_frame,
    text="NPK PHONSKA I",
    font=("Segoe UI", 18, "bold"),
    bg="#0f0f0f",
    fg="#FFD700"   # emas
)
title2.pack()

# ==========================
# VIDEO CENTER
# ==========================

video_container = tk.Frame(content_frame, bg="#111111")
video_container.grid(row=0, column=1,sticky="nsew", padx=(0,20), pady=20)

video_label = tk.Label(video_container, bg="black")
video_label.pack(fill="both", expand=True)
# ==========================
# FPS LABEL DI BAWAH KIRI VIDEO
# ==========================
fps_overlay = tk.Label(
    video_container,
    text="FPS: 0",
    font=("Arial", 10, "bold"),
    bg="#111111",
    fg="lime"
)
fps_overlay.place(relx=0.01, rely=0.95, anchor="sw")

# STATUS RTSP LABEL
rtsp_status_label = tk.Label(
    video_container,
    text="RTSP: -",
    font=("Arial", 10, "bold"),
    bg="#111111",
    fg="white"
)
rtsp_status_label.place(relx=0.01, rely=0.99, anchor="sw")

# =========================================================
# CAMERA INIT
# =========================================================
print("Opening RTSP stream...")

# menggunakan substream H264 channel 102 (lebih stabil)
os.environ["OPENCV_FFMPEG_HW_ACCELERATION"] = "d3d11va"
os.environ["OPENCV_FFMPEG_HW_DEVICE"] = "0"
cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
# buffer 0 = anti delay
cap.set(cv2.CAP_PROP_BUFFERSIZE, 0)
cap.set(cv2.CAP_PROP_FPS, 30)
cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 2000)
cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 2000)

root.mainloop()
cap.release()
cv2.destroyAllWindows()
