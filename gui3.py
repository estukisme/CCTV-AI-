import cv2
import numpy as np
import tensorflow as tf
import torch
from ultralytics import YOLO
import os
import tkinter as tk
from tkinter import Label, ttk
from tkinter.scrolledtext import ScrolledText
from PIL import Image, ImageTk
from datetime import datetime
import threading
import winsound
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import openpyxl



# =========================================================
# GPU SETUP
# =========================================================
torch.backends.cudnn.benchmark = True
DEVICE = "cuda"

print("Torch:", torch.__version__)
print("CUDA:", torch.cuda.is_available())
print("GPU:", torch.cuda.get_device_name(0))

gpus = tf.config.list_physical_devices('GPU')
if gpus:
    for gpu in gpus:
        tf.config.experimental.set_memory_growth(gpu, True)

# =========================================================
# GLOBAL
# =========================================================
running = False
processing = False
alarm_active = False
buzzer_running = False
current_frame = None
heatmap_enabled = True
last_time = time.time()
fps = 0
export_enabled = True
current_shift_name = None
excel_lock = threading.Lock()
ROI_X1 = 500
ROI_Y1 = 200
ROI_X2 = 1400
ROI_Y2 = 850
current_shift_alarm_file = None
current_shift_monitor_file = None
last_monitor_minute = None
rtsp_url = "rtsp://admin:Phonska1bisa@192.168.110.3:554/Streaming/Channels/101?rtsp_transport=tcp"


# =========================================================
# LOAD MODELS
# =========================================================
yolo = YOLO("yolov8n.pt")
yolo.to(DEVICE)
yolo.fuse()

cnn = tf.keras.models.load_model(
    "cnn/model_cnn_v2.h5",
    compile=False
)

cnn_labels = ["merah", "tidak_merah"]
torch.set_float32_matmul_precision('high')

# =========================================================
# SHIFT
# =========================================================
def get_shift(now):
    hour = now.hour
    if 7 <= hour < 15:
        return "Shift 1 (07:00-15:00)"
    elif 15 <= hour < 23:
        return "Shift 2 (15:00-23:00)"
    else:
        return "Shift 3 (23:00-07:00)"
    
def get_alarm_filename(now):
    shift = get_shift(now)
    date_str = now.strftime("%Y-%m-%d")
    return f"{date_str}_{shift.replace(' ','_').replace(':','')}_ALARM.xlsx"

def get_monitor_filename(now):
    shift = get_shift(now)
    date_str = now.strftime("%Y-%m-%d")
    return f"{date_str}_{shift.replace(' ','_').replace(':','')}_MONITORING.xlsx"
    
# =========================================================
# EXPORT EXCEL
# =========================================================
def append_to_excel(data,filename):


    if filename is None:
        return

    with excel_lock:

        os.makedirs("export", exist_ok=True)
        filepath = os.path.join("export", filename)

        if os.path.exists(filepath):
            try:
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
            except:
                os.remove(filepath)
                wb = Workbook()
                ws = wb.active
                ws.title = "Log Produk"
                ws.append(["Foto Produk", "Tanggal", "Jam", "Shift", "Warna Produk"])
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log Produk"
            ws.append(["Foto Produk", "Tanggal", "Jam", "Shift", "Warna Produk"])
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 18

        row = ws.max_row + 1

        ws.cell(row=row, column=2, value=data["date"])
        ws.cell(row=row, column=3, value=data["time"])
        ws.cell(row=row, column=4, value=data["shift"])
        ws.cell(row=row, column=5, value=data["status"])

        if os.path.exists(data["image_path"]):
            img = XLImage(data["image_path"])
            img.width = 120
            img.height = 120
            ws.add_image(img, f"A{row}")
            ws.row_dimensions[row].height = 95

        temp_path = filepath + ".tmp"
        wb.save(temp_path)
        wb.close()
        # Replace file lama secara aman
        os.replace(temp_path, filepath)
    
# =========================================================
# BUZZER
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOUND_FILE = os.path.join(BASE_DIR, "alarm.wav")
def buzzer_loop():
    global buzzer_running

    while buzzer_running:

        # Play sampai selesai (blocking, bukan async)
        winsound.PlaySound(SOUND_FILE, winsound.SND_FILENAME)

        # Setelah suara selesai, jeda 3 detik sebelum mengulang
        time.sleep(3)   

def start_buzzer():
    global buzzer_running
    if not buzzer_running:
        buzzer_running = True
        threading.Thread(target=buzzer_loop, daemon=True).start()

def stop_buzzer():
    global buzzer_running
    buzzer_running = False
    winsound.PlaySound(None, winsound.SND_PURGE)

def resize_with_aspect_ratio(image, width=None, height=None):
    h, w = image.shape[:2]

    if width is None and height is None:
        return image

    if width is not None:
        ratio = width / w
        new_dim = (width, int(h * ratio))
    else:
        ratio = height / h
        new_dim = (int(w * ratio), height)

    return cv2.resize(image, new_dim)

   
# =========================================================
# CNN PREPROCESS
# =========================================================
def preprocess_cnn(image_bgr):
    rgb = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB)
    resized = cv2.resize(rgb, (64, 64))
    return resized.astype("float32") / 255.0

# =========================================================
# DETECTION
# =========================================================
def detect_objects(frame):
    global alarm_active, processing, current_frame
    global current_shift_name
    global current_shift_alarm_file, current_shift_monitor_file
    global last_monitor_minute, export_enabled

    detected_color = "-"
    roi = frame[ROI_Y1:ROI_Y2, ROI_X1:ROI_X2]

    if roi.size != 0:

        inp = preprocess_cnn(roi).reshape(1, 64, 64, 3)
        pred = cnn.predict(inp, verbose=0)[0]
        confidence_merah = pred[0]     # index 0 = merah
        confidence_tidak = pred[1]     # index 1 = tidak merah

        if confidence_merah > 0.75:
            detected_color = "merah"
        else:
            detected_color = "tidak_merah"

        color_box = (0,0,255) if detected_color=="merah" else (0,255,255)

        cv2.rectangle(frame,(ROI_X1, ROI_Y1),(ROI_X2, ROI_Y2),color_box,3)
        cv2.putText(frame,detected_color.upper(),
                    (ROI_X1, ROI_Y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,color_box,3)

        if export_enabled:

            now = datetime.now()
            shift_name = get_shift(now)

            # ===== SHIFT CHANGE =====
            if current_shift_name != shift_name:
                current_shift_name = shift_name
                current_shift_alarm_file = get_alarm_filename(now)
                current_shift_monitor_file = get_monitor_filename(now)
                last_monitor_minute = None

            # ===== ALARM FILE =====
            if detected_color == "tidak_merah":

                os.makedirs("snapshot_alarm", exist_ok=True)
                img_name = now.strftime("%Y%m%d_%H%M%S") + "_ALARM.jpg"
                img_path = os.path.join("snapshot_alarm", img_name)
                cv2.imwrite(img_path, roi)

                data = {
                    "date": now.strftime("%Y-%m-%d"),
                    "time": now.strftime("%H:%M:%S"),
                    "shift": shift_name,
                    "status": detected_color,
                    "image_path": img_path
                }

                append_to_excel(data, current_shift_alarm_file)

            # ===== MONITORING FILE (00 & 30) =====
            if now.minute in [0, 30] and now.second <= 5:

                if last_monitor_minute != now.minute:

                    last_monitor_minute = now.minute

                    os.makedirs("snapshot_monitoring", exist_ok=True)
                    img_name = now.strftime("%Y%m%d_%H%M%S") + ".jpg"
                    img_path = os.path.join("snapshot_monitoring", img_name)
                    cv2.imwrite(img_path, roi)

                    data = {
                        "date": now.strftime("%Y-%m-%d"),
                        "time": now.strftime("%H:%M:%S"),
                        "shift": shift_name,
                        "status": detected_color,
                        "image_path": img_path
                    }

                    append_to_excel(data, current_shift_monitor_file)

    # ===== BUZZER CONTROL =====
    if detected_color == "tidak_merah":
        if not alarm_active:
            alarm_active = True
            start_buzzer()
    else:
        if alarm_active:
            alarm_active = False
            stop_buzzer()

    current_frame = frame
    processing = False


# =========================================================
# HEATMAP
# =========================================================

def create_heatmap(frame):
    # Ambil channel merah
    red_channel = frame[:, :, 2]

    # Normalize
    heat = cv2.normalize(red_channel, None, 0, 255, cv2.NORM_MINMAX)

    # Convert jadi colormap
    heatmap = cv2.applyColorMap(heat, cv2.COLORMAP_JET)

    # Perkecil heatmap
    heatmap_small = cv2.resize(heatmap, (220, 150))

    return heatmap_small

def toggle_heatmap():
    global heatmap_enabled
    heatmap_enabled = not heatmap_enabled
    
    if heatmap_enabled:
        btn_heatmap.config(text="Heatmap: ON")
    else:
        btn_heatmap.config(text="Heatmap: OFF")

def force_reconnect():
    global cap

    print("FORCE RECONNECTING RTSP ...")

    try:
        cap.release()
    except:
        pass

    time.sleep(0.3)

    cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
# =========================================================
# MAIN LOOP (STABIL)
# =========================================================
def main_loop():
    global running, processing, current_frame
    global last_time, fps
    global cap

    if not running:
        return

    ret, frame = cap.read()
    
    if not ret or frame is None:
        print("RTSP lost... reconnecting")
        force_reconnect()
        root.after(300, main_loop)
        return

    # Jalankan deteksi async
    if not processing:
        processing = True
        threading.Thread(
            target=detect_objects,
            args=(frame.copy(),),
            daemon=True
        ).start()

    # Frame display
    display = current_frame if current_frame is not None else frame

    root.update_idletasks()
    container_w = video_container.winfo_width()
    container_h = video_container.winfo_height()

    if container_w > 10 and container_h > 10:
        h, w = display.shape[:2]
        scale = min(container_w / w, container_h / h)
        new_w = int(w * scale)
        new_h = int(h * scale)
        resized = cv2.resize(display, (new_w, new_h))
        # Buat canvas hitam sebesar container
        canvas = np.zeros((container_h, container_w, 3), dtype=np.uint8)

        # Hitung offset tengah
        x_offset = (container_w - new_w) // 2
        y_offset = 0

        # Tempel video di tengah
        canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized

        display = canvas

    # ==========================
    # FPS CALCULATION
    # ==========================
    current_time = time.time()
    fps = 1 / (current_time - last_time)
    last_time = current_time

    fps_int = int(fps)

    if fps_int < 10:
        color = "red"
    elif 10 <= fps_int <= 15:
        color = "yellow"
    else:
        color = "lime"

    fps_label.config(text=f"FPS: {fps_int}", fg=color)

    # ==========================
    # HEATMAP SECTION
    # ==========================
    if heatmap_enabled:
        heatmap = create_heatmap(display)
        heat_rgb = cv2.cvtColor(heatmap, cv2.COLOR_BGR2RGB)
        heat_img = ImageTk.PhotoImage(Image.fromarray(heat_rgb))

        heatmap_label.imgtk = heat_img
        heatmap_label.configure(image=heat_img)
    else:
        blank = np.zeros((150, 220, 3), dtype=np.uint8)
        blank_rgb = cv2.cvtColor(blank, cv2.COLOR_BGR2RGB)
        blank_img = ImageTk.PhotoImage(Image.fromarray(blank_rgb))

        heatmap_label.imgtk = blank_img
        heatmap_label.configure(image=blank_img)
        
    # ==========================
    # VIDEO DISPLAY
    # ==========================
    rgb = cv2.cvtColor(display, cv2.COLOR_BGR2RGB)
    img = ImageTk.PhotoImage(Image.fromarray(rgb))

    video_label.imgtk = img
    video_label.configure(image=img)
    root.after(15, main_loop)

# =========================================================
# SNAPSHOT
# =========================================================
def snapshot():
    if current_frame is None:
        return
    fname = datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"
    cv2.imwrite(os.path.join("snapshot", fname), current_frame)

def toggle_export():
    global export_enabled
    export_enabled = not export_enabled
    btn_export.config(text=f"Export: {'ON' if export_enabled else 'OFF'}")
# =========================================================
# CONTROL
# =========================================================
def start_camera():
    global running
    running = True
    dashboard_frame.pack_forget()
    camera_frame.pack(fill="both", expand=True)
    main_loop()

def stop_camera():
    global running
    running = False
    stop_buzzer()

def back_to_dashboard():
    global running
    running = False
    stop_buzzer()
    camera_frame.pack_forget()
    dashboard_frame.pack(fill="both", expand=True)

# =========================================================
# GUI
# =========================================================
root = tk.Tk()
root.title("Hybrid YOLO + CNN RTX 4070 Ti STABLE")
root.state("zoomed")
root.configure(bg="#1e1e1e")

style = ttk.Style()
style.theme_use("default")

# ==========================
# DASHBOARD
# ==========================
dashboard_frame = tk.Frame(root, bg="#1e1e1e")

tk.Label(dashboard_frame, text="Dashboard",
         fg="white", bg="#1e1e1e",
         font=("Arial", 26, "bold")).pack(pady=50)

ttk.Button(dashboard_frame, text="Mulai Kamera",
           command=start_camera).pack(pady=20)

ttk.Button(dashboard_frame, text="Keluar",
           command=root.quit).pack(pady=20)

dashboard_frame.pack(fill="both", expand=True)

# ==========================
# CAMERA FRAME
# ==========================

camera_frame = tk.Frame(root, bg="#1e1e1e")

# ==========================
# TOP BAR MONITORING
# ==========================
top_bar = tk.Frame(camera_frame, bg="#0f0f0f", height=60)
top_bar.pack(fill="x")

title_label = tk.Label(top_bar,
                       text="CCTV AI",
                       font=("Arial", 20, "bold"),
                       bg="#0f0f0f",
                       fg="white")

title_label.pack(side="left", padx=30)
middle_title = tk.Label(
    top_bar,
    text="DEP. OPERASI IIA - NPK PHONSKA 1",
    font=("Segoe UI", 22, "bold"),
    bg="#0f0f0f",
    fg="#FFD700"
)
middle_title.pack(side="top", pady=5)


# FPS
fps_label = tk.Label(
    top_bar,
    text="FPS: 0",
    font=("Arial", 16, "bold"),
    bg="#0f0f0f",
    fg="lime"
)
fps_label.pack(side="right", padx=20)

#logo
logo_img = Image.open("logo.png")
# Perbesar dengan menjaga rasio
base_height = 60
h_percent = base_height / float(logo_img.size[1])
new_width = int(float(logo_img.size[0]) * h_percent)
logo_img = logo_img.resize((new_width, base_height), Image.LANCZOS)
logo_tk = ImageTk.PhotoImage(logo_img)

logo_label = tk.Label(top_bar,
                      image=logo_tk,
                      bg="#0f0f0f")
logo_label.image = logo_tk
logo_label.pack(side="right", padx=20)

#logo 2
logo_img = Image.open("logo KIK new 2.png")
# Perbesar dengan menjaga rasio
base_height = 60
h_percent = base_height / float(logo_img.size[1])
new_width = int(float(logo_img.size[0]) * h_percent)
logo_img = logo_img.resize((new_width, base_height), Image.LANCZOS)
logo_tk = ImageTk.PhotoImage(logo_img)

logo_label = tk.Label(top_bar,
                      image=logo_tk,
                      bg="#fffdfd")
logo_label.image = logo_tk
logo_label.pack(side="right", padx=20)

#logo 3
logo_img = Image.open("POG.jpg")
# Perbesar dengan menjaga rasio
base_height = 60
h_percent = base_height / float(logo_img.size[1])
new_width = int(float(logo_img.size[0]) * h_percent)
logo_img = logo_img.resize((new_width, base_height), Image.LANCZOS)
logo_tk = ImageTk.PhotoImage(logo_img)

logo_label = tk.Label(top_bar,
                      image=logo_tk,
                      bg="#fffdfd")
logo_label.image = logo_tk
logo_label.pack(side="right", padx=20)
# ==========================
# CONTENT FRAME (CENTER MODE)
# ==========================
content_frame = tk.Frame(camera_frame, bg="#1e1e1e")
content_frame.pack(fill="both", expand=True)


# 1 kolom saja
content_frame.columnconfigure(0, weight=0)
content_frame.columnconfigure(1, weight=1)  # video fleksibel


# 3 row
content_frame.rowconfigure(0, weight=1)  # JUDUL kecil


# ==========================
# PANEL KIRI (BUTTON)
# ==========================
btn_panel = tk.Frame(content_frame, bg="#1e1e1e")
btn_panel.grid(row=0, column=0, sticky="ns", padx=20, pady=20)

ttk.Button(btn_panel, text="Stop",
           command=stop_camera).pack(fill="x", pady=(0, 15))

ttk.Button(btn_panel, text="Snapshot",
           command=snapshot).pack(fill="x", pady=(0, 15))

btn_heatmap = ttk.Button(btn_panel,
                         text="Heatmap: ON",
                         command=toggle_heatmap)
btn_heatmap.pack(fill="x", pady=(0, 15))

btn_export = ttk.Button(btn_panel,
                        text="Export: ON",
                        command=toggle_export)
btn_export.pack(fill="x", pady=(0, 15))

ttk.Button(btn_panel, text="Kembali",
           command=back_to_dashboard).pack(fill="x")

heatmap_label = tk.Label(btn_panel, bg="black")
heatmap_label.pack(pady=(30, 0))

# ==========================
# VIDEO CENTER
# ==========================

video_container = tk.Frame(content_frame, bg="#111111")
video_container.grid(row=0, column=1,sticky="nsew", padx=(0,20), pady=20)

video_label = tk.Label(video_container, bg="black")
video_label.pack(fill="both", expand=True)

# =========================================================
# CAMERA INIT
# =========================================================
cap = cv2.VideoCapture(rtsp_url, cv2.CAP_FFMPEG)
cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

root.mainloop()
cap.release()
cv2.destroyAllWindows()
